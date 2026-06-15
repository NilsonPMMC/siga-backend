import os
import datetime
import openpyxl

from django.core.management.base import BaseCommand, CommandError

from atendimentos.models import Municipe


class Command(BaseCommand):
    help = (
        "Gera relatório Excel de munícipes filtrando por IDs de match "
        "vindos de uma planilha (ex.: candidatos_vereadores_match_*.xlsx). "
        "Não altera nenhum dado do banco."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--planilha-match",
            type=str,
            required=True,
            help="Caminho da planilha com coluna 'Match ID'.",
        )
        parser.add_argument(
            "--arquivo",
            type=str,
            default=None,
            help="Nome do arquivo Excel de saída.",
        )
        parser.add_argument(
            "--ordenar-por",
            type=str,
            default="nome",
            choices=["nome", "orgao"],
            help="Ordenação do relatório final.",
        )

    def _extrair_ids_match(self, caminho_planilha):
        if not os.path.exists(caminho_planilha):
            raise CommandError(f"Planilha não encontrada: {caminho_planilha}")

        wb = openpyxl.load_workbook(caminho_planilha, read_only=True, data_only=True)
        try:
            ws = wb["Matches Candidatos"] if "Matches Candidatos" in wb.sheetnames else wb.active
            headers = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            try:
                idx_status = headers.index("Status")
                idx_match_id = headers.index("Match ID")
            except ValueError as exc:
                raise CommandError(
                    "Colunas obrigatórias não encontradas na planilha. "
                    "Esperado: 'Status' e 'Match ID'."
                ) from exc

            ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                status = str(row[idx_status]).strip().upper() if row[idx_status] is not None else ""
                raw_id = row[idx_match_id]
                if status == "ENCONTRADO" and raw_id not in (None, ""):
                    try:
                        ids.add(int(raw_id))
                    except (TypeError, ValueError):
                        continue
            return sorted(ids)
        finally:
            wb.close()

    def handle(self, *args, **options):
        caminho_planilha = options["planilha_match"]
        ordenar_por = options["ordenar_por"]
        nome_arquivo = options["arquivo"]

        match_ids = self._extrair_ids_match(caminho_planilha)
        if not match_ids:
            raise CommandError("Nenhum Match ID válido encontrado na planilha.")

        self.stdout.write(self.style.SUCCESS(f"IDs de match carregados: {len(match_ids)}"))

        order_fields = ["orgao", "nome_completo"] if ordenar_por == "orgao" else ["nome_completo"]
        queryset = (
            Municipe.objects.filter(id__in=match_ids)
            .prefetch_related("perfis__categoria", "contas")
            .order_by(*order_fields)
        )

        if not nome_arquivo:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
            nome_arquivo = f"relatorio_municipes_por_match_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contatos Match"
        ws.append([
            "Match ID",
            "Nome Completo",
            "Nome de Guerra",
            "CPF",
            "Data de Nascimento",
            "Email Principal",
            "Telefone Principal",
            "Cargo",
            "Órgão",
            "Categoria",
            "Contas Vinculadas",
        ])

        for m in queryset:
            email_principal = ""
            if isinstance(m.emails, list) and m.emails:
                first_email = m.emails[0]
                email_principal = first_email.get("email", "") if isinstance(first_email, dict) else str(first_email)

            telefone_principal = ""
            if isinstance(m.telefones, list) and m.telefones:
                first_fone = m.telefones[0]
                telefone_principal = first_fone.get("numero", "") if isinstance(first_fone, dict) else str(first_fone)

            categorias = sorted({p.categoria.nome for p in m.perfis.all() if p.categoria})
            contas_vinculadas = ", ".join([c.nome for c in m.contas.all()])

            ws.append([
                m.id,
                m.nome_completo,
                m.nome_de_guerra or "",
                m.cpf or "",
                m.data_nascimento.strftime("%d/%m/%Y") if m.data_nascimento else "",
                email_principal,
                telefone_principal,
                m.cargo or "",
                m.orgao or "",
                ", ".join(categorias),
                contas_vinculadas,
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        wb.save(nome_arquivo)
        self.stdout.write(self.style.SUCCESS(f"Relatório salvo em: {os.path.abspath(nome_arquivo)}"))
        self.stdout.write(self.style.SUCCESS(f"Total exportado: {queryset.count()}"))

