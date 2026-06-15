"""
Comando Django para identificar candidatos a vereadores 2024 no banco de dados.

Uso:
    python manage.py match_candidatos_vereadores [--aplicar] [--categoria-id=8] [--conta-id=1]

Flags:
    --aplicar: Aplica as mudanças (atualiza/cria registros). Sem essa flag, roda em dry-run.
    --categoria-id: ID da categoria a ser atribuída (padrão: 8 - VEREADORES)
    --conta-id: ID da conta para vincular (padrão: 1,2)
    --threshold: Score mínimo de similaridade para match (padrão: 85)
    --export: Exporta relatório Excel com resultados
"""

import csv
import unicodedata
from pathlib import Path
from django.core.management.base import BaseCommand
from django.db.models import Q
from atendimentos.models import Municipe, PerfilMunicipe, CategoriaContato, Conta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def normalizar_nome(nome):
    """
    Normaliza nome para comparação:
    - Remove acentos
    - Converte para uppercase
    - Remove caracteres especiais extras
    """
    if not nome:
        return ""
    
    # Remove acentos
    nome = unicodedata.normalize('NFD', nome)
    nome = ''.join(char for char in nome if unicodedata.category(char) != 'Mn')
    
    # Uppercase e limpa
    nome = nome.upper().strip()
    
    # Remove múltiplos espaços
    nome = ' '.join(nome.split())
    
    return nome


def calcular_similaridade(str1, str2):
    """
    Calcula similaridade entre duas strings usando Levenshtein distance.
    Retorna score de 0 a 100.
    """
    try:
        from fuzzywuzzy import fuzz
        return fuzz.ratio(str1, str2)
    except ImportError:
        # Fallback simples se fuzzywuzzy não estiver instalado
        if str1 == str2:
            return 100
        if str1 in str2 or str2 in str1:
            return 80
        return 0


def buscar_match_municipe(nome_candidato, threshold=85):
    """
    Busca o melhor match para um candidato no banco de dados.
    
    Estratégia otimizada:
    1. Busca exata por nome completo (normalizado)
    2. Busca exata por nome de guerra
    3. Busca por partes do nome (ILIKE)
    4. Fuzzy matching apenas se não encontrar nas etapas anteriores
    
    Returns:
        tuple: (municipe, score, campo_match) ou (None, 0, None)
    """
    nome_normalizado = normalizar_nome(nome_candidato)
    
    # Extrai primeiras palavras para busca inteligente
    palavras = nome_normalizado.split()
    primeira_palavra = palavras[0] if palavras else ""
    ultima_palavra = palavras[-1] if len(palavras) > 1 else ""
    
    # 1. Busca exata (case-insensitive)
    match_exato = Municipe.objects.filter(
        Q(nome_completo__iexact=nome_candidato) |
        Q(nome_de_guerra__iexact=nome_candidato)
    ).first()
    
    if match_exato:
        campo = 'nome_completo' if match_exato.nome_completo.upper() == nome_normalizado else 'nome_de_guerra'
        return match_exato, 100, campo
    
    # 2. Busca por partes do nome (ILIKE - PostgreSQL/MySQL)
    # Busca por primeira e última palavra para reduzir candidatos
    candidatos_db = Municipe.objects.filter(
        Q(nome_completo__icontains=primeira_palavra) |
        Q(nome_de_guerra__icontains=primeira_palavra) |
        Q(nome_completo__icontains=ultima_palavra) |
        Q(nome_de_guerra__icontains=ultima_palavra)
    ).only('id', 'nome_completo', 'nome_de_guerra')[:200]  # Limita para performance
    
    if not candidatos_db:
        return None, 0, None
    
    # 3. Fuzzy matching nos candidatos filtrados
    melhor_match = None
    melhor_score = 0
    campo_match = None
    
    for municipe in candidatos_db:
        # Testa nome_completo
        if municipe.nome_completo:
            nome_completo_norm = normalizar_nome(municipe.nome_completo)
            score_completo = calcular_similaridade(nome_normalizado, nome_completo_norm)
            
            if score_completo > melhor_score:
                melhor_score = score_completo
                melhor_match = municipe
                campo_match = 'nome_completo'
        
        # Testa nome_de_guerra
        if municipe.nome_de_guerra:
            nome_guerra_norm = normalizar_nome(municipe.nome_de_guerra)
            score_guerra = calcular_similaridade(nome_normalizado, nome_guerra_norm)
            
            if score_guerra > melhor_score:
                melhor_score = score_guerra
                melhor_match = municipe
                campo_match = 'nome_de_guerra'
    
    # Só retorna se atingiu o threshold
    if melhor_score >= threshold:
        return melhor_match, melhor_score, campo_match
    
    return None, 0, None


class Command(BaseCommand):
    help = 'Identifica candidatos a vereadores 2024 no banco de dados usando fuzzy matching'

    def add_arguments(self, parser):
        parser.add_argument(
            '--aplicar',
            action='store_true',
            help='Aplica mudanças (atualiza/cria registros). Sem essa flag, roda em dry-run.',
        )
        parser.add_argument(
            '--categoria-id',
            type=int,
            default=8,
            help='ID da categoria a ser atribuída (padrão: 8 - VEREADORES)',
        )
        parser.add_argument(
            '--conta-id',
            type=str,
            default='1,2',
            help='IDs das contas para vincular, separados por vírgula (padrão: 1,2)',
        )
        parser.add_argument(
            '--threshold',
            type=int,
            default=85,
            help='Score mínimo de similaridade para match (0-100, padrão: 85)',
        )
        parser.add_argument(
            '--export',
            type=str,
            default=None,
            help='Caminho para exportar relatório Excel',
        )

    def handle(self, *args, **options):
        aplicar = options['aplicar']
        categoria_id = options['categoria_id']
        conta_ids = [int(x.strip()) for x in options['conta_id'].split(',')]
        threshold = options['threshold']
        export_path = options['export']
        
        # Modo
        modo = 'APLICAR' if aplicar else 'DRY-RUN'
        
        self.stdout.write(self.style.WARNING('=' * 80))
        self.stdout.write(self.style.WARNING(f'IDENTIFICAÇÃO DE CANDIDATOS A VEREADORES 2024 - Modo: {modo}'))
        self.stdout.write(self.style.WARNING('=' * 80))
        self.stdout.write(f'Categoria ID: {categoria_id}')
        self.stdout.write(f'Contas IDs: {conta_ids}')
        self.stdout.write(f'Threshold: {threshold}%')
        self.stdout.write('')
        
        # Valida categoria
        try:
            categoria = CategoriaContato.objects.get(id=categoria_id)
            self.stdout.write(self.style.SUCCESS(f'✓ Categoria encontrada: {categoria.nome}'))
        except CategoriaContato.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'✗ Categoria ID {categoria_id} não encontrada!'))
            return
        
        # Valida contas
        contas = Conta.objects.filter(id__in=conta_ids)
        if not contas.exists():
            self.stdout.write(self.style.ERROR(f'✗ Nenhuma conta encontrada com IDs {conta_ids}!'))
            return
        self.stdout.write(self.style.SUCCESS(f'✓ Contas encontradas: {", ".join([c.nome for c in contas])}'))
        self.stdout.write('')
        
        # Lê CSV de candidatos
        csv_path = Path('/var/www/gabinete/docs/planilha_candidatos_vereadores_2024.csv')
        if not csv_path.exists():
            self.stdout.write(self.style.ERROR(f'✗ Arquivo não encontrado: {csv_path}'))
            return
        
        # Detecta encoding
        import chardet
        with open(csv_path, 'rb') as f:
            raw_data = f.read()
            detected = chardet.detect(raw_data)
            encoding = detected['encoding']
            self.stdout.write(f'Encoding detectado: {encoding} (confiança: {detected["confidence"]*100:.1f}%)')
        
        candidatos = []
        with open(csv_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f)
            for row in reader:
                nome = row.get('Nome', '').strip()
                if nome:
                    candidatos.append(nome)
        
        self.stdout.write(self.style.SUCCESS(f'✓ {len(candidatos)} candidatos carregados do CSV'))
        self.stdout.write('')
        
        # Processa cada candidato
        resultados = []
        encontrados = 0
        nao_encontrados = 0
        atualizados = 0
        criados = 0
        
        self.stdout.write(self.style.WARNING('Processando candidatos...'))
        self.stdout.write('')
        
        for idx, nome_candidato in enumerate(candidatos, 1):
            self.stdout.write(f'[{idx}/{len(candidatos)}] {nome_candidato}')
            
            # Busca match
            municipe, score, campo = buscar_match_municipe(nome_candidato, threshold)
            
            if municipe:
                encontrados += 1
                self.stdout.write(self.style.SUCCESS(
                    f'  ✓ MATCH: {municipe.nome_completo} (ID: {municipe.id}, '
                    f'campo: {campo}, score: {score}%)'
                ))
                
                # Verifica se já tem perfil com a categoria
                perfis_existentes = PerfilMunicipe.objects.filter(
                    municipe=municipe,
                    conta__in=contas,
                    categoria=categoria
                )
                
                if perfis_existentes.exists():
                    self.stdout.write('    → Já possui perfil com esta categoria')
                    status_acao = 'JA_EXISTE'
                else:
                    if aplicar:
                        # Cria perfil para cada conta
                        for conta in contas:
                            perfil, created = PerfilMunicipe.objects.get_or_create(
                                municipe=municipe,
                                conta=conta,
                                defaults={
                                    'categoria': categoria,
                                    'cargo': 'CANDIDATO A VEREADOR 2024',
                                    'ativo': True
                                }
                            )
                            if created:
                                criados += 1
                                self.stdout.write(self.style.SUCCESS(
                                    f'    ✓ Perfil criado para conta {conta.nome}'
                                ))
                            else:
                                # Atualiza categoria se necessário
                                if perfil.categoria != categoria:
                                    perfil.categoria = categoria
                                    perfil.cargo = 'CANDIDATO A VEREADOR 2024'
                                    perfil.save()
                                    atualizados += 1
                                    self.stdout.write(self.style.SUCCESS(
                                        f'    ✓ Perfil atualizado para conta {conta.nome}'
                                    ))
                        status_acao = 'ATUALIZADO' if atualizados > 0 else 'CRIADO'
                    else:
                        self.stdout.write(self.style.WARNING(
                            '    → [DRY-RUN] Perfil seria criado/atualizado'
                        ))
                        status_acao = 'SERIA_ATUALIZADO'
                
                resultados.append({
                    'candidato': nome_candidato,
                    'match_id': municipe.id,
                    'match_nome': municipe.nome_completo,
                    'match_apelido': municipe.nome_de_guerra or '',
                    'campo_match': campo,
                    'score': score,
                    'status': 'ENCONTRADO',
                    'acao': status_acao
                })
            else:
                nao_encontrados += 1
                self.stdout.write(self.style.ERROR(f'  ✗ NÃO ENCONTRADO (nenhum match acima de {threshold}%)'))
                
                resultados.append({
                    'candidato': nome_candidato,
                    'match_id': None,
                    'match_nome': '',
                    'match_apelido': '',
                    'campo_match': '',
                    'score': 0,
                    'status': 'NAO_ENCONTRADO',
                    'acao': 'NENHUMA'
                })
            
            self.stdout.write('')
        
        # Resumo
        self.stdout.write(self.style.WARNING('-' * 80))
        self.stdout.write(self.style.WARNING('RESUMO:'))
        self.stdout.write(self.style.SUCCESS(f'  ✓ Encontrados: {encontrados}'))
        self.stdout.write(self.style.ERROR(f'  ✗ Não encontrados: {nao_encontrados}'))
        if aplicar:
            self.stdout.write(self.style.SUCCESS(f'  ✓ Perfis criados: {criados}'))
            self.stdout.write(self.style.SUCCESS(f'  ✓ Perfis atualizados: {atualizados}'))
        self.stdout.write(f'  Taxa de sucesso: {(encontrados/len(candidatos)*100):.1f}%')
        self.stdout.write(self.style.WARNING('-' * 80))
        
        # Exporta relatório Excel
        if export_path or not aplicar:
            if not export_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                export_path = f'/var/www/gabinete/tmp/candidatos_vereadores_match_{timestamp}.xlsx'
            
            self.exportar_excel(resultados, export_path, modo, threshold)
            self.stdout.write(self.style.SUCCESS(f'\n✓ Relatório exportado: {export_path}'))

    def exportar_excel(self, resultados, caminho, modo, threshold):
        """Exporta resultados para Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Matches Candidatos"
        
        # Cabeçalho
        headers = [
            'Candidato', 'Status', 'Match ID', 'Nome Completo (BD)', 
            'Nome de Guerra (BD)', 'Campo Match', 'Score %', 'Ação'
        ]
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dados
        for row_idx, resultado in enumerate(resultados, 2):
            ws.cell(row=row_idx, column=1, value=resultado['candidato'])
            ws.cell(row=row_idx, column=2, value=resultado['status'])
            ws.cell(row=row_idx, column=3, value=resultado['match_id'])
            ws.cell(row=row_idx, column=4, value=resultado['match_nome'])
            ws.cell(row=row_idx, column=5, value=resultado['match_apelido'])
            ws.cell(row=row_idx, column=6, value=resultado['campo_match'])
            ws.cell(row=row_idx, column=7, value=resultado['score'])
            ws.cell(row=row_idx, column=8, value=resultado['acao'])
            
            # Cor da linha baseada no status
            if resultado['status'] == 'ENCONTRADO':
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = fill
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20
        
        # Adiciona resumo
        ws_summary = wb.create_sheet("Resumo")
        ws_summary.append(['Modo', modo])
        ws_summary.append(['Threshold', f'{threshold}%'])
        ws_summary.append([''])
        ws_summary.append(['Total de Candidatos', len(resultados)])
        ws_summary.append(['Encontrados', sum(1 for r in resultados if r['status'] == 'ENCONTRADO')])
        ws_summary.append(['Não Encontrados', sum(1 for r in resultados if r['status'] == 'NAO_ENCONTRADO')])
        ws_summary.append(['Taxa de Sucesso', f"{(sum(1 for r in resultados if r['status'] == 'ENCONTRADO')/len(resultados)*100):.1f}%"])
        
        wb.save(caminho)
